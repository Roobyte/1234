import vk_api
import pyodbc
import requests
import openpyxl
import gspread
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from datetime import datetime, timezone
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api import VkUpload

# Кнопки
keyboard = VkKeyboard(one_time=False)
keyboard.add_button('Стандартные команды', color=VkKeyboardColor.NEGATIVE)
keyboard.add_button('Погода Барнаул', color=VkKeyboardColor.NEGATIVE)
keyboard.add_line()
keyboard.add_button('Статус', color=VkKeyboardColor.POSITIVE)
keyboard.add_button('Добыть репутации',
                    color=VkKeyboardColor.SECONDARY)
keyboard.add_line()
keyboard.add_button('Установить никнейм', color=VkKeyboardColor.POSITIVE)
keyboard.add_button('Показать статус по NickName',
                    color=VkKeyboardColor.POSITIVE)
keyboard.add_line()
keyboard.add_button('Режим записи в таблицу exel',
                    color=VkKeyboardColor.POSITIVE)
# функция, которая отправляет сообщение пользователю


def write_message(sender, message, st=0):
    if st != 0:
        authorize.method('messages.send', {
            'user_id': sender, 'sticker_id': st, 'random_id': get_random_id()})
    authorize.method('messages.send', {
        'user_id': sender, 'message': message, 'random_id': get_random_id(), 'keyboard': keyboard.get_keyboard()})
##################################### Функции с бд, проверка на наличеи и добавление пользователей#####################################
# функция, которая проверяет наличие пользователя


def check_and_insert_id(id):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = f"SELECT id_users FROM users WHERE id_users = {id}"
    cursor.execute(query)
    result = cursor.fetchone()

    if result:
        conn.close()
        return
    query = f"INSERT INTO users (id_users) VALUES ('{id}')"
    cursor.execute(query)
    conn.commit()
    conn.close()

# функия, которая заполняет список пользователей


def list_user_id(list_user):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = "SELECT id_users FROM users"
    cursor.execute(query)
    result = cursor.fetchall()
    for i in result:
        list_user.append(i[0])
    cursor.close()
    conn.close()

# Заполнение статистики пользователей


def user_statusPlus(list_user):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = "SELECT * FROM user_stats"
    cursor.execute(query)
    result = cursor.fetchall()
    for i in result:
        list_user.append({'id_users': i[0], 'job_title': i[1], 'Reputation': i[2],
                         'number_of_messages': i[3], 'game_currency': i[4], 'Nickname': i[5]})
    cursor.close()
    conn.close()

# учёт статистики пользователей


def user_statusMinus(sender):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = "SELECT * FROM user_stats"
    cursor.execute(query)
    index = 0
    result = cursor.fetchall()
    while (True):
        if (int(sender)) == (int(result[index][0])):
            break
        else:
            index += 1
        if index == len(result):
            break
    number_of_messages = result[index][3]
    number_of_messages = number_of_messages+1
    game_currency = result[index][4]
    game_currency += 1
    query = f"UPDATE user_stats SET job_title='Появилась', number_of_messages = {
        number_of_messages}, game_currency = {game_currency} WHERE id_users = {sender}"
    cursor.execute(query)
    conn.commit()
    cursor.close()
    conn.close()


def info_user_stat(sender, reseived_message=""):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    if len(reseived_message) > 0:
        query = f"SELECT * FROM user_stats WHERE Nickname = '{
            reseived_message}'"
    else:
        query = f"SELECT * FROM user_stats WHERE id_users = {sender}"
    cursor.execute(query)
    result = cursor.fetchone()
    write_message(sender, f"✦ Должность: {result[1]}\n ↑↑ Репутация: {result[2]}\n ⇆ Количество сообщений: {
                  result[3]}\n ₽ Игровой валюты: {result[4]}\n Ⓡ Никнейм: {result[5]}")
    cursor.close()
    conn.close()
    return result


def install_nickname(sender, nickname):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = f"UPDATE user_stats SET Nickname = ? WHERE id_users = {sender}"
    cursor.execute(query, nickname)
    conn.commit()
    cursor.close()
    conn.close()


def seacrh_user_stat(nickname):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = f"SELECT id_users, Nickname FROM user_stats WHERE Nickname = ?"
    cursor.execute(query, nickname)
    result = cursor.fetchall()
    cursor.close()
    conn.close()
    if len(result) > 0 and len(result) < 2:
        return result[0][1]
    else:
        return ""


def check_and_insert_id_users_stat(id):
    conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    query = f"SELECT id_users FROM user_stats WHERE id_users = {id}"
    cursor.execute(query)
    result = cursor.fetchone()
    DEFAULT_JOB_TITLE = "Нет должности"
    DEFAULT_REPUTATION = 0
    DEFAULT_NUMBER_OF_MESSAGES = 0
    DEFAULT_GAME_CURRENCY = 0
    DEFAULT_NICKNAME = f"{id}"
    if result:
        conn.close()
        return
    query = f"INSERT INTO user_stats (id_users, job_title, Reputation, number_of_messages, game_currency, Nickname) VALUES (?, ?, ?, ?, ?, ?)"
    cursor.execute(query, id, DEFAULT_JOB_TITLE, DEFAULT_REPUTATION,
                   DEFAULT_NUMBER_OF_MESSAGES, DEFAULT_GAME_CURRENCY, DEFAULT_NICKNAME)
    conn.commit()
    conn.close()

##################################### Функции с бд, проверка на наличеи и добавление пользователей#####################################


def write_to_excel(filename, sheetname, cell, data, sender):
    # открытие файла
    wb = openpyxl.load_workbook(filename)
    # выбор листа
    print(data, sender)
    write_message(sender, "1")
    sheet = wb[sheetname]
    # получение значения ячейки
    value = sheet[cell].value
    # проверка на значение
    write_message(sender, "2")
    if not value:
        write_message(sender, "3")
        # записываем ну ты понял
        sheet[cell] = data
    else:
        write_message(sender, "4")
        row = sheet.max_row + 1
        while sheet.cell(row=row, column=sheet[cell].column).value:
            row += 1
        sheet.cell(row=row, column=sheet[cell].column).value = data
        write_message(sender, "5")
    wb.save(filename)
    write_message(sender, "Успешно")


################################# Прогноз погоды Барнаул #################################
# получение погоды Барнаул
# Запрос текущей погоды

############################# EXEL добавление и тд#######################

# Проверка ячейки на пустоту


# Найти первую пустую ячейку в столбце column_index

############################# EXEL добавление и тд#######################
def request_current_weather(sender):
    city_id = 1510853
    appid = "e5ad8d7039ee5c2c7cf99d7178e818e3"
    try:
        res = requests.get("http://api.openweathermap.org/data/2.5/weather",
                           params={'id': city_id, 'units': 'metric', 'lang': 'ru', 'APPID': appid})
        data = res.json()
        weather_description = data['weather'][0]['description']
        temperature = data['main']['temp']
        min_temperature = data['main']['temp_min']
        max_temperature = data['main']['temp_max']
        feels_like = data['main']['feels_like']
        wind = data['wind']['speed']
        # Отправка сообщений с использованием функции write_message
        write_message(sender, f"Погода: {weather_description}")
        write_message(sender, f"Температура: {temperature}°C")
        write_message(sender, f"Минимальная температура: {min_temperature}°C")
        write_message(sender, f"Максимальная температура: {max_temperature}°C")
        write_message(sender, f"Ощущается как: {feels_like}°C")
        write_message(sender, f"Скорость ветра: {wind} м/с")
    except Exception as e:
        print("Exception (weather):", e)
        pass


def get_wind_direction(deg):
    l = ['С ', 'СВ', ' В', 'ЮВ', 'Ю ', 'ЮЗ', ' З', 'СЗ']
    for i in range(0, 8):
        step = 45.
        min = i*step - 45/2.
        max = i*step + 45/2.
        if i == 0 and deg > 360-45/2.:
            deg = deg - 360
        if deg >= min and deg <= max:
            res = l[i]
            break
    return res

# Прогноз


def request_forecast(sender):
    city_id = 1510853
    appid = "e5ad8d7039ee5c2c7cf99d7178e818e3"
    count = 0
    try:
        res = requests.get("http://api.openweathermap.org/data/2.5/forecast",
                           params={'id': city_id, 'units': 'metric', 'lang': 'ru', 'APPID': appid})
        data = res.json()

        current_date = ""
        forecast = ""

        for i in data['list']:
            date = i['dt_txt'][:10]
            if date != current_date:  # Если новая дата, показываем прогноз на весь день
                if current_date != "":  # Пропускаем первую итерацию, так как текущая дата еще не инициализирована
                    write_message(sender, f"Дата: {current_date}")
                    write_message(sender, f"Прогноз на весь день: {forecast}")
                    write_message(sender, "________________________")
                current_date = date
                forecast = ""

            temperature = '{0:>+3.0f}'.format(i['main']['temp'])
            weather = i['weather'][0]['description']
            forecast += f"Время: {i['dt_txt'][11:16]
                                  } | Температура: {temperature}°C | Погода: {weather}\n"

        # Отправляем последний прогноз на весь день
        if current_date != "":
            write_message(sender, f"Дата: {current_date}")
            write_message(sender, f"Прогноз на весь день: {forecast}")

    except Exception as e:
        print("Exception (forecast):", e)
        pass


################################ Прогноз погоды Барнаул #################################

##################### Расписание###########################


# Инициализация веб-драйвера
def schedule():
    options = Options()
    options.use_insecure_ssl = True

    # Используйте путь к драйверу Microsoft Edge на вашем компьютере
    service = Service(
        executable_path='./msedgedriver.exe')
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)

    driver.get('http://www.asiec.ru/ras/#')
    but = driver.find_element(By.ID, 'details-button')
    but.click()
    link = driver.find_element(By.ID, 'proceed-link')
    link.click()
    # details-button
    # proceed-link
    date_input = driver.find_element(By.ID, 'calendar2')
    date_input.clear()
    date_input.send_keys('21-10-2023')

    submit_button = driver.find_element(
        By.CSS_SELECTOR, 'input[type="submit"]')
    submit_button.click()

    driver.implicitly_wait(10)

    content = driver.find_element(By.ID, 'content').text
    print(content)
    open('paspis.txt', 'a', encoding='utf-8').write(content)
    driver.quit()


schedule()
##################### Расписание###########################
token = "vk1.a.MBhu7p7myjCxozC4kCVaUTnSwZDoDFmUuMHGEjOAvedPpk361-hZnfmFpBc26Jct5rwtZBF3AzoKHOlwbpjU_FNIPi8zub9gUSEN9dvhOMobpCO2_JgtkXEcZPUX732ZtQ39RsnqXH97AWV5O7aIxci7tyjW34PeLWmeWWEHjhmrkN9BgKuSyTB2jTF7V8nOk5la3UlN8vIDakJ89OlKjA"
authorize = vk_api.VkApi(token=token)
questions = {"Собака это млекопитающие?": 'да', 'Люди это млекопитающие?': 'да',
             'Утконос это млекопитающие?': 'да', 'а килька это млекопитающие?': 'нет'}
upload = VkUpload(authorize)
list_user = []
list_status_user = []
list_user_id(list_user)
user_statusPlus(list_status_user)
# логика бота для моей группы, фигня не рабочая ) кек
# longpoll = VkBotLongPoll(authorize, 'club222923184')

longpoll = VkLongPoll(authorize)
while True:
    try:
        for event in longpoll.listen():
            if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
                reseived_message = event.text
                sender = event.user_id
                dataa = datetime.now(timezone.utc).strftime('%d.%m.%Y %H: %M')
                data = (event.user_id, reseived_message, dataa)
                # Работа с бд - сообщения
                open('bot_messeges.txt', 'a', encoding='utf-8').write(
                    str(event.user_id) + ':' + reseived_message + " |дата написания:" + dataa + '\n')
                conn_str = 'Driver={SQL Server};Server=DESKTOP-VMD04M2\SQLEXPRESS;Database=VK_BOT_base;Trusted_Connection=yes;'
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()
                sql_query = f"INSERT INTO message (id, message, data) VALUES (?, ?, ?)"
                check_and_insert_id(sender)
                user_statusPlus(list_status_user)
                check_and_insert_id_users_stat(sender)
                user_statusMinus(sender)
                cursor.execute(sql_query, data)
                conn.commit()
                conn.close()
                # Добыть репутации
                ####
                if reseived_message == "Режим записи в таблицу exel":
                    write_message(
                        sender, "Режим включин вставляй ссылку")
                    for event in longpoll.listen():
                        if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
                            ss = event.text
                            sender = event.user_id
                            write_to_excel('LinkIns.xlsx', 'user',
                                           'B2',  ss, sender)
                            write_message(
                                sender, "Для выхода из режима напишете \"Выход\" ")
                            if ss == "Выход":
                                break

                ####
                if reseived_message == 'Добыть репутацию':
                    write_message(
                        sender, "Добрый день, теперь ты можешь добыть репутацию(неработает)")
                # Статус пользователя
                if reseived_message == 'Статус':
                    info_user_stat(sender)
                # Прогноз погода Барнаул на 5 дней
                if reseived_message == 'Прогноз':
                    request_forecast(sender)
                # Кнопка Погода Барнаул
                if reseived_message == 'Погода Барнаул':
                    request_current_weather(sender)
                # Сообщение 'Для Ильи!'
                if reseived_message == 'Для Ильи!' and sender == 421345334:
                    write_message(
                        sender, "Удивлён, что ты в этой группе!", 81923)
                    write_message(
                        sender, "Ладно, это чат бот, который ничего не умеет кроме как стандартный команд, которые сделаны в if @_@", 84981)
                # Сообщение 'Для тебя Вадим!'
                if reseived_message == 'Для тебя Вадим!' and sender == 302700118:
                    write_message(
                        sender, "Даже не пытайся найти пасхалки :))", 84981)
                # Сообщение 'Для тебя Даниил!'
                if reseived_message == 'Для тебя Даниил!' and sender == 405499977:
                    write_message(
                        sender, "Ну хз, спс за помощь с тестами бота!", 84981)
                # Поиск статус по NickName
                if reseived_message == "Показать статус по NickName":
                    write_message(
                        sender, "Пожалуйста введите никнейм пользователя, которого хотите проверить")
                    for event in longpoll.listen():
                        if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
                            reseived_message = event.text
                            if reseived_message == seacrh_user_stat(reseived_message):
                                info_user_stat(sender, reseived_message)
                                break
                            else:
                                write_message(
                                    sender, "Пользовтель с таким никнейном не найден")
                                break
                # Сообщение Привет
                if reseived_message == 'Привет':
                    write_message(sender, "Меня зовут bot puvel", 81923)
                # Сообщение Пока
                if reseived_message == 'Пока':
                    write_message(sender, "Бай бай")
                if reseived_message == 'Установить никнейм':
                    write_message(
                        sender, "Введите свой никнейм")
                    for event in longpoll.listen():
                        if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
                            reseived_message = event.text
                            if reseived_message != seacrh_user_stat(reseived_message):
                                install_nickname(sender, reseived_message)
                                write_message(
                                    sender, "Никнейм успешно установлен")
                                break
                            else:
                                write_message(
                                    sender, "Неудача, такой никнейм уже существует, или введены некорректные данные")
                                break
                # Нажатие на кнопку 'Стандартные команды'
                if reseived_message == 'Стандартные команды':
                    if sender == 405499977:
                        write_message(
                            sender, "Стандартные команды: Привет, Пока, Для тебя Даниил!")
                    elif sender == 302700118:
                        write_message(
                            sender, "Стандартные команды: Привет, Пока, Для тебя Вадим!")
                    elif sender == 421345334:
                        write_message(
                            sender, "Стандартные команды: Привет, Пока, Для Ильи!")
                    else:
                        write_message(
                            sender, "Стандартные команды: Привет, Пока,Прогноз ##(на 5 дней вперёд, Барнаул)##")
                # длина листа
                if reseived_message == 'длина':
                    write_message(
                        sender, f"{len(list_user)}", 81923)

                # Рассылка
                if reseived_message == 'Рассылка' and sender == 466412337:
                    for user_id in list_user:
                        write_message(
                            user_id, f"Это рассылка! Получите погоду!", 81923)
                        request_forecast(user_id)
    except:
        pass
